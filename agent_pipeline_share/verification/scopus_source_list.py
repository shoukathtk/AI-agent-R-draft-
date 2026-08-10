"""
Scopus indexing check via the official Scopus Source List.

Elsevier publishes a free spreadsheet of every ISSN currently indexed in
Scopus, updated a few times a year:
  https://www.elsevier.com/products/scopus/scopus-source-list
(Look for "Scopus Source List" / "Source Titles List" download — an .xlsx file.)

This is the authoritative free method: no API key, no rate limit, and it's
literally the list Elsevier itself indexes from. The tradeoff is it's a
point-in-time snapshot, not live — re-download every few months.

There is no equivalent free file or API for Web of Science. Clarivate's
Master Journal List (https://mjl.clarivate.com) is web-search-only; checking
WoS status has to stay a manual step unless you have a paid WoS API license
(see scopus_api.py's WOS_NOTE for details).
"""

import csv
from pathlib import Path
from functools import lru_cache

try:
    import openpyxl  # only needed if you point this at the raw .xlsx
except ImportError:
    openpyxl = None


def _normalize_issn(issn: str) -> str:
    return issn.replace("-", "").replace(" ", "").upper().strip()


@lru_cache(maxsize=1)
def _load_index(source_list_path: str) -> dict:
    """
    Builds an ISSN -> record lookup from the Scopus Source List.
    Accepts either a .csv (recommended: export the xlsx to CSV once) or a
    .xlsx directly if openpyxl is installed.

    Expected columns (Elsevier's actual export has these, names vary slightly
    by release — adjust COLUMN_MAP below if a column isn't found):
      Source Title, Print-ISSN, E-ISSN, CiteScore, Subject area, Active/Inactive
    """
    path = Path(source_list_path)
    if not path.exists():
        raise FileNotFoundError(
            f"Scopus Source List not found at {source_list_path}. "
            "Download it from https://www.elsevier.com/products/scopus/scopus-source-list "
            "and point SCOPUS_SOURCE_LIST_PATH at it (CSV recommended)."
        )

    index = {}
    if path.suffix.lower() == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                _index_row(row, index)
    elif path.suffix.lower() in (".xlsx", ".xlsm"):
        if openpyxl is None:
            raise RuntimeError("pip install openpyxl --break-system-packages, or export the file to CSV instead.")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            _index_row(row_dict, index)
    else:
        raise ValueError("Source list must be .csv or .xlsx")

    return index


COLUMN_MAP = {
    "title": ["Source Title", "Title", "Source title"],
    "print_issn": ["Print-ISSN", "Print ISSN", "ISSN"],
    "e_issn": ["E-ISSN", "E ISSN", "eISSN"],
    "citescore": ["CiteScore", "CiteScore 2023", "CiteScore 2024"],
    "subject_area": ["Subject area", "Scopus Sub-Subject Area", "All Science Journal Classification Codes (ASJC)"],
    "active": ["Active or Inactive", "Active/Inactive", "Status"],
}


def _get_col(row: dict, key: str):
    for candidate in COLUMN_MAP[key]:
        if candidate in row and row[candidate] not in (None, ""):
            return row[candidate]
    return None


def _index_row(row: dict, index: dict) -> None:
    title = _get_col(row, "title")
    record = {
        "title": title,
        "citescore": _get_col(row, "citescore"),
        "subject_area": _get_col(row, "subject_area"),
        "active": _get_col(row, "active"),
    }
    for issn_key in ("print_issn", "e_issn"):
        raw = _get_col(row, issn_key)
        if raw:
            index[_normalize_issn(str(raw))] = record


def check_scopus_indexing(issn_list: list[str], source_list_path: str) -> dict:
    """
    issn_list: ISSNs pulled from Crossref for a given citation (print and/or e-ISSN).
    Returns a verification result dict.
    """
    try:
        index = _load_index(source_list_path)
    except (FileNotFoundError, RuntimeError, ValueError) as e:
        return {"status": "SOURCE_LIST_UNAVAILABLE", "error": str(e)}

    for issn in issn_list:
        norm = _normalize_issn(issn)
        if norm in index:
            record = index[norm]
            active = record.get("active")
            is_active = active is None or str(active).strip().lower().startswith("active")
            return {
                "status": "SCOPUS_CONFIRMED" if is_active else "SCOPUS_INACTIVE",
                "matched_issn": issn,
                "journal_title": record.get("title"),
                "citescore": record.get("citescore"),
                "subject_area": record.get("subject_area"),
            }
    return {"status": "NOT_IN_SCOPUS_LIST", "checked_issns": issn_list}
